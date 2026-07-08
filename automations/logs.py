import os
import queue
import shlex
import threading
from datetime import datetime
import paramiko
from dotenv import load_dotenv

load_dotenv()

ELASTICSEARCHS = {
    "Elasticsearch 1": os.getenv("ELASTIC_1_IP", ""),
    "Elasticsearch 2": os.getenv("ELASTIC_2_IP", ""),
    "Elasticsearch 3": os.getenv("ELASTIC_3_IP", ""),
}

LIFERAYS = {
    "Liferay 1": os.getenv("LIFERAY_1_IP", ""),
    "Liferay 2": os.getenv("LIFERAY_2_IP", ""),
}

USUARIO = os.getenv("LIFERAY_USUARIO")
SENHA   = os.getenv("LIFERAY_SENHA")

ARQUIVO_LOG_ELASTIC = "/var/log/elasticsearch/LiferayElasticsearchCluster.log"
BASE_LIFERAY = "/home/liferay/logs/liferay.{data}"

PADROES = {
    '<log4j:message><![CDATA[org.elasticsearch.ElasticsearchStatusException:':
        "ElasticSearchException",
    '<log4j:message><![CDATA[{code="404", msg="Not Found",':
        "Notfound",
    '<log4j:message><![CDATA[java.io.IOException: Broken pipe]]></log4j:message>':
        "Broken Pipe",
    '<log4j:message><![CDATA[java.io.IOException: Connection reset by peer]]></log4j:message>':
        "Connection Reset By Peer",
    'ldap.internal.exportimport.LDAPUserImporterImpl':
        "LDAPUserImporterImpl",
    '<log4j:message><![CDATA[Problem accessing LDAP server]]></log4j:message>':
        "Problem accessing LDAP server",
    '(ldap.internal.authenticator.LDAPAuth)':
        "LDAPAuth",
    '<log4j:message><![CDATA[Skip /o/js/resolved-module/frontend-js-metal-web$metal':
        "Skip /frontend-js-metal-web$metal",
    '<log4j:message><![CDATA[No theme found for specified theme id mtportal_WAR_mtportaltheme.':
        "No theme found for mtportal_WAR",
    '<log4j:message><![CDATA[{code="404"':
        'Error {code="404"}',
    '<log4j:message><![CDATA[SAX Security Manager':
        "SAX Security Manager",
    'uri=/mt-portal-theme/images/ajax-loader.gif':
        "Ajax-Loader",
    'com.liferay.change.tracking.internal.servlet.filter.CTCollectionPreviewFilter':
        "CTCollectionPreviewFilter",
    'FTL stack trace':
        "FTL stack trace",
    'log4j:throwable><![CDATA[com.liferay.document.library.kernel.exception.NoSuchFileEntryException':
        "NoSuchFileEntryException",
    'InvalidRepositoryIdException':
        "InvalidRepositoryIdException",
    '[PaginationLimitFilter] Applying limit: bot UA detected':
        "PaginationLimitFilter bot UA",
}


def _ssh_client(ip):
    ssh = paramiko.SSHClient()
    ssh.load_system_host_keys()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ip, username=USUARIO, password=SENHA, timeout=20)
    return ssh


def _exec_cmd(ssh, cmd):
    _, stdout, stderr = ssh.exec_command(cmd)
    out = stdout.read().decode(errors="ignore").strip()
    err = stderr.read().decode(errors="ignore").strip()
    return out, err


def _obter_tamanho_elastic(nome, ip):
    try:
        ssh = _ssh_client(ip)
        out, err = _exec_cmd(ssh, f"stat -c%s {ARQUIVO_LOG_ELASTIC}")
        ssh.close()
        if err:
            return None, f"{nome}: erro ao obter tamanho ({err})"
        return int(out), None
    except Exception as e:
        return None, f"{nome}: falha de conexao ({e})"


def _obter_tamanho_liferay(nome, ip, data_iso):
    try:
        ssh = _ssh_client(ip)
        arq = BASE_LIFERAY.format(data=data_iso) + ".log"
        out, err = _exec_cmd(ssh, f"stat -c%s '{arq}'")
        ssh.close()
        if err:
            return None, f"{nome}: arquivo LOG nao encontrado para {data_iso}"
        return int(out), None
    except Exception as e:
        return None, f"{nome}: falha na etapa tamanho LOG ({e})"


def _grep_padroes(ssh, xml):
    # Todos os 16 greps em um único exec_command.
    # grep -c sempre imprime o contador (inclusive "0") quando o arquivo existe,
    # então NÃO usamos "|| echo 0" — isso causaria linha dupla para padrões com
    # 0 matches (grep sai com código 1 mas já imprimiu "0"), deslocando o mapeamento.
    parts = [
        f"grep -F -c {shlex.quote(busca)} {shlex.quote(xml)} 2>/dev/null"
        for busca in PADROES
    ]
    cmd = "; ".join(parts)
    out, _ = _exec_cmd(ssh, cmd)
    lines = out.splitlines()
    totais = dict.fromkeys(PADROES.values(), 0)
    for i, exibicao in enumerate(PADROES.values()):
        if i < len(lines):
            try:
                totais[exibicao] = int(lines[i].strip())
            except ValueError:
                totais[exibicao] = 0
    return totais


def _grep_ftl_templates(ssh, xml):
    cmd = f"grep -oP '\\[in template \"[^\"]*\"' '{xml}' | sort | uniq -c | sort -rn"
    out, _ = _exec_cmd(ssh, cmd)
    templates = {}
    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(None, 1)
        if len(parts) == 2:
            try:
                templates[parts[1]] = int(parts[0])
            except ValueError:
                pass
    return templates


def _contar_padroes(nome, ip, data_iso):
    """Retorna (resultado, erro). Faz tudo em UMA conexão SSH: XML + tamanho do .log."""
    try:
        ssh = _ssh_client(ip)
        xml  = BASE_LIFERAY.format(data=data_iso) + ".xml"
        log_ = BASE_LIFERAY.format(data=data_iso) + ".log"

        # Verifica existência do XML e busca tamanho do .log na mesma chamada
        check_cmd = (
            f"test -f {shlex.quote(xml)} && echo XML_OK || echo XML_MISSING; "
            f"stat -c%s {shlex.quote(log_)} 2>/dev/null || echo LOG_MISSING"
        )
        out, _ = _exec_cmd(ssh, check_cmd)
        linhas = out.splitlines()
        xml_ok   = len(linhas) > 0 and linhas[0].strip() == "XML_OK"
        log_raw  = linhas[1].strip() if len(linhas) > 1 else "LOG_MISSING"
        tam_log  = int(log_raw) if log_raw.isdigit() else None

        if not xml_ok:
            ssh.close()
            return None, f"{nome}: arquivo XML nao encontrado para {data_iso}"

        totais = _grep_padroes(ssh, xml)
        ftl_templates = {}
        if totais.get("FTL stack trace", 0) > 0:
            ftl_templates = _grep_ftl_templates(ssh, xml)
        ssh.close()
        return {"totais": totais, "ftl_templates": ftl_templates, "tam_log": tam_log}, None
    except Exception as e:
        return None, f"{nome}: falha na analise XML ({e})"


def _exportar_excel(result: dict, output_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        HDR_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        ALT_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
        HDR_FONT = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="CCCCCC")
        BRD = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Sheet 1: Contagem de erros por data
        ws = wb.active
        ws.title = "Contagem de Erros"
        erros = list(PADROES.values())
        datas = sorted(result["por_data"].keys())
        ws.append(["Erro"] + datas)
        for cell in ws[1]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BRD
        for i, erro in enumerate(erros, 2):
            row = [erro] + [result["por_data"][d]["contagem"].get(erro, 0) for d in datas]
            ws.append(row)
            for j, cell in enumerate(ws[i]):
                cell.border = BRD
                if i % 2 == 0:
                    cell.fill = ALT_FILL
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value or "")) for c in col) + 2

        # Sheet 2: Tamanhos dos logs
        ws2 = wb.create_sheet("Tamanhos dos Logs")
        ws2.append(["Servidor", "Data", "Tamanho (MB)"])
        for cell in ws2[1]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BRD
        r = 2
        for data_iso, info in result["por_data"].items():
            for nome, tam in info["tamanhos_logs"].items():
                mb = round(tam / 1024 / 1024, 2) if tam else None
                ws2.append([nome, data_iso, mb])
                for cell in ws2[r]:
                    cell.border = BRD
                r += 1
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 22

        # Sheet 3: Elasticsearch bytes
        ws3 = wb.create_sheet("Elasticsearch")
        ws3.append(["Servidor", "Bytes", "MB"])
        for cell in ws3[1]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BRD
        for nome, val in result["elastic_bytes"].items():
            mb = round(val / 1024 / 1024, 2) if val else None
            ws3.append([nome, val, mb])
        for col in ws3.columns:
            ws3.column_dimensions[col[0].column_letter].width = 22

        wb.save(output_path)
    except Exception as e:
        raise RuntimeError(f"Erro ao exportar Excel: {e}") from e


def run_logs(datas: list, out_q: queue.Queue, stop_ev: threading.Event,
             servidores_elastic: list | None = None,
             servidores_liferay: list | None = None):
    def log(msg, tag="normal"):
        out_q.put((tag, msg + "\n"))

    elastic_filtro = servidores_elastic or list(ELASTICSEARCHS.keys())
    liferay_filtro = servidores_liferay or list(LIFERAYS.keys())
    # 1 passo por elastic + 1 passo por liferay por data (tamanho agora vem junto)
    total_steps = len(elastic_filtro) + len(datas) * len(liferay_filtro)
    step = 0

    log("  Iniciando coleta de logs via SSH...", "info")
    log(f"  Elasticsearch: {', '.join(elastic_filtro)}", "dim")
    log(f"  Liferay: {', '.join(liferay_filtro)}", "dim")
    log("")

    # Tamanhos Elasticsearch
    log("  -- Elasticsearch " + "-" * 42, "dim")
    elastic_bytes = {}
    for nome in elastic_filtro:
        if stop_ev.is_set():
            break
        ip = ELASTICSEARCHS.get(nome)
        if not ip:
            continue
        log(f"  -> Conectando {nome} ({ip})...", "progress")
        tamanho, erro = _obter_tamanho_elastic(nome, ip)
        if erro:
            log(f"  [!] {erro}", "warning")
        else:
            log(f"  OK {nome}: {tamanho:,} bytes", "success")
        elastic_bytes[nome] = tamanho
        step += 1
        out_q.put(("PROGRESS", step / max(total_steps, 1)))

    result_por_data = {}

    for data in datas:
        if stop_ev.is_set():
            break

        data_iso = data.strftime("%Y-%m-%d")
        data_br  = data.strftime("%d/%m/%Y")

        log(f"\n  {'─' * 54}", "dim")
        log(f"  Data: {data_br}", "title")
        log(f"  {'─' * 54}", "dim")

        consolidado     = dict.fromkeys(PADROES.values(), 0)
        ftl_consolidado = {}
        tamanhos_logs   = {}

        for nome in liferay_filtro:
            if stop_ev.is_set():
                break
            ip = LIFERAYS.get(nome)
            if not ip:
                continue

            log(f"\n  -> Analisando {nome} ({data_iso})...", "info")
            contagens, erro = _contar_padroes(nome, ip, data_iso)
            if erro:
                log(f"  [!] {erro}", "warning")
                tamanhos_logs[nome] = None
            else:
                for k, v in contagens["totais"].items():
                    consolidado[k] += v
                for tmpl, cnt in contagens["ftl_templates"].items():
                    ftl_consolidado[tmpl] = ftl_consolidado.get(tmpl, 0) + cnt
                # tam_log vem embutido — sem conexao SSH extra
                tamanhos_logs[nome] = contagens.get("tam_log")
            step += 1
            out_q.put(("PROGRESS", step / max(total_steps, 1)))

        log("")
        for i, (nome_erro, qty) in enumerate(consolidado.items(), 1):
            if qty == 0:
                tag = "dim"
            elif qty < 5:
                tag = "success"
            elif qty < 20:
                tag = "warning"
            else:
                tag = "error"
            bar = "X" * min(qty, 30)
            log(f"  {i:02d}. {nome_erro:<52} {qty:>5}   {bar}", tag)

        log("")
        for nome in liferay_filtro:
            tam = tamanhos_logs.get(nome)
            if tam is None:
                log(f"  LOG {nome}: arquivo nao encontrado", "warning")
            else:
                log(f"  LOG {nome}: {tam / 1024 / 1024:.1f} MB", "info")

        result_por_data[data_iso] = {
            "contagem":      consolidado,
            "tamanhos_logs": tamanhos_logs,
            "ftl_templates": ftl_consolidado,
        }

    if stop_ev.is_set():
        log("\n  [!] Execucao interrompida.", "warning")
        out_q.put(("DONE", None))
        return

    log(f"\n  {'─' * 54}", "dim")
    log("  Tamanho em bytes — Elasticsearch:", "title")
    log("")
    for nome, valor in elastic_bytes.items():
        if valor is None:
            log(f"  {nome}: erro ao obter", "warning")
        else:
            log(f"  {nome}: {valor:,} bytes", "info")

    # FTL stack trace consolidado
    ftl_geral = {}
    for info in result_por_data.values():
        for tmpl, cnt in info["ftl_templates"].items():
            ftl_geral[tmpl] = ftl_geral.get(tmpl, 0) + cnt

    if ftl_geral:
        log(f"\n  {'─' * 54}", "dim")
        log("  Erros de FTL stack trace:", "title")
        log("")
        for tmpl, cnt in sorted(ftl_geral.items(), key=lambda x: -x[1]):
            log(f"  {tmpl}  ({cnt}x)", "warning")

    log("\n  Coleta concluida com sucesso!", "success")
    out_q.put(("PROGRESS", 1.0))

    result_data = {
        "elastic_bytes": elastic_bytes,
        "por_data":      result_por_data,
    }
    out_q.put(("RESULT_DATA", result_data))

    total_erros = sum(
        sum(info["contagem"].values()) for info in result_por_data.values()
    )
    out_q.put(("HISTORY", {
        "modulo": "logs",
        "status": "sucesso",
        "detalhes": f"{len(datas)} data(s) | {total_erros} ocorrencias no total",
    }))
    out_q.put(("DONE", None))
