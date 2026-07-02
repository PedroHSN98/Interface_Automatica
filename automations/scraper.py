import os
import json
import hashlib
import queue
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urljoin, urlparse


TAMANHO_MIN = 40960     # 40 KB
TAMANHO_MAX = 2097152   # 2 MB
_HASHES_FILE = ".hashes.json"


def _extrair_portal(url: str) -> str:
    try:
        partes = urlparse(url).path.split("/")
        if "web" in partes:
            idx = partes.index("web")
            if len(partes) > idx + 1 and partes[idx + 1]:
                return partes[idx + 1].upper()
        return "GERAL"
    except Exception:
        return "ERRO_URL"


def _load_hashes(output_folder: str) -> set:
    path = os.path.join(output_folder, _HASHES_FILE)
    try:
        with open(path, "r", encoding="utf-8") as f:
            return set(json.load(f))
    except (FileNotFoundError, json.JSONDecodeError):
        return set()


def _save_hashes(output_folder: str, hashes: set) -> None:
    path = os.path.join(output_folder, _HASHES_FILE)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(list(hashes), f)


def _baixar_imagem(url_img: str, url_origem: str, output_folder: str,
                   sites_pesados: set, hashes_conhecidos: set) -> tuple[bool, bool, str | None]:
    """Retorna (baixou, fora_padrao, hash_novo)."""
    try:
        import requests
        head = requests.head(url_img, timeout=5, allow_redirects=True)
        tam = int(head.headers.get("Content-Length", 0))
        if tam > TAMANHO_MAX:
            sites_pesados.add(url_origem)
            return False, True, None
        if TAMANHO_MIN <= tam <= TAMANHO_MAX:
            res = requests.get(url_img, timeout=10)
            if res.status_code == 200:
                img_hash = hashlib.md5(res.content).hexdigest()
                if img_hash in hashes_conhecidos:
                    return False, False, None  # duplicata
                nome = f"img_{img_hash}.jpg"
                with open(os.path.join(output_folder, nome), "wb") as f:
                    f.write(res.content)
                return True, False, img_hash
        return False, False, None
    except Exception:
        return False, False, None


def run_scraper(links_path: str, output_folder: str, out_q: queue.Queue, stop_ev: threading.Event):
    def log(msg, tag="normal"):
        out_q.put((tag, msg + "\n"))

    try:
        import requests
        from bs4 import BeautifulSoup
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError as e:
        log(f"  [ERRO] Dependencia nao instalada: {e}", "error")
        log("  Execute: pip install requests beautifulsoup4 pandas openpyxl", "warning")
        out_q.put(("DONE", None))
        return

    ARQUIVO_DADOS = os.path.join(output_folder, "registro_noticias.xlsx")
    sites_pesados: set = set()

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    if not os.path.exists(links_path):
        log(f"  [ERRO] Arquivo nao encontrado: {links_path}", "error")
        out_q.put(("DONE", None))
        return

    with open(links_path, "r") as f:
        urls = [l.strip() for l in f if l.strip()]

    if not urls:
        log("  [ERRO] Nenhuma URL encontrada no arquivo de links.", "error")
        out_q.put(("DONE", None))
        return

    # Carrega hashes de imagens ja baixadas para deteccao de duplicatas
    hashes_conhecidos = _load_hashes(output_folder)
    hashes_novos: set = set()

    log(f"  Iniciando varredura em {len(urls)} link(s)...", "info")
    log(f"  Pasta de saida: {output_folder}", "info")
    log(f"  Filtro de tamanho: {TAMANHO_MIN // 1024}KB - {TAMANHO_MAX // 1024 // 1024}MB", "info")
    log(f"  Hashes conhecidos (duplicatas a ignorar): {len(hashes_conhecidos)}", "dim")
    log("")

    relatorio = []
    for i, url in enumerate(urls, 1):
        if stop_ev.is_set():
            log("\n  [!] Execucao interrompida.", "warning")
            break

        nome_portal = _extrair_portal(url)
        pct = i / len(urls)
        barra = "X" * int(30 * pct) + "." * (30 - int(30 * pct))
        out_q.put(("progress", f"  [{barra}] {i}/{len(urls)}  {nome_portal}\n"))
        out_q.put(("PROGRESS", pct))

        try:
            res = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
            sopa = BeautifulSoup(res.text, "html.parser")
            imgs = [urljoin(url, img.get("src")) for img in sopa.find_all("img") if img.get("src")]

            todos_hashes = hashes_conhecidos | hashes_novos

            def _baixar(img_url, _url=url, _th=todos_hashes):
                return _baixar_imagem(img_url, _url, output_folder, sites_pesados, _th)

            with ThreadPoolExecutor(max_workers=5) as ex:
                resultados = list(ex.map(_baixar, imgs))

            sucessos = sum(1 for r in resultados if r[0])
            fora     = sum(1 for r in resultados if r[1])

            # Registra hashes novos
            for baixou, _, h in resultados:
                if baixou and h:
                    hashes_novos.add(h)

            tag = "success" if sucessos > 0 else "dim"
            log(f"      {nome_portal:<30} encontradas={len(imgs):>3}  baixadas={sucessos:>3}  fora={fora:>3}", tag)

            relatorio.append({
                "Data":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Portal":      nome_portal,
                "URL":         url,
                "Encontradas": len(imgs),
                "Baixadas":    sucessos,
                "Fora_Padrao": fora,
            })
        except Exception as e:
            log(f"      [ERRO] {url[:60]}... -> {e}", "error")

    # Persiste hashes acumulados
    _save_hashes(output_folder, hashes_conhecidos | hashes_novos)

    if relatorio:
        df = pd.DataFrame(relatorio)
        if os.path.exists(ARQUIVO_DADOS):
            df_old = pd.read_excel(ARQUIVO_DADOS)
            df = pd.concat([df_old, df], ignore_index=True)
        df.to_excel(ARQUIVO_DADOS, index=False, engine="openpyxl")

        wb = load_workbook(ARQUIVO_DADOS)
        ws = wb.active
        fill_hdr = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill_hdr
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            vals = [str(c.value) for c in col if c.value]
            if vals:
                ws.column_dimensions[col[0].column_letter].width = max(len(v) for v in vals) + 2
        wb.save(ARQUIVO_DADOS)

        total_baixado = sum(r["Baixadas"] for r in relatorio)
        maximo        = max(r["Baixadas"] for r in relatorio)

        log("")
        log("  " + "=" * 56, "dim")
        log("  VARREDURA FINALIZADA!", "title")
        log(f"  Total de imagens baixadas: {total_baixado}", "success")
        log(f"  Recorde em um unico portal: {maximo}", "success")
        log(f"  Duplicatas ignoradas: {len(hashes_novos)} novas / {len(hashes_conhecidos)} ja conhecidas", "dim")
        log(f"  Planilha salva: {ARQUIVO_DADOS}", "info")

        if sites_pesados:
            log("")
            log("  ALERTA - Imagens acima de 2MB detectadas em:", "warning")
            for site in sites_pesados:
                log(f"    -> {site}", "warning")

        log("  " + "=" * 56, "dim")

    out_q.put(("PROGRESS", 1.0))
    out_q.put(("DONE_SCRAPER", output_folder))
    out_q.put(("HISTORY", {
        "modulo": "scraper",
        "status": "sucesso",
        "detalhes": f"{len(relatorio)} portais | {sum(r['Baixadas'] for r in relatorio)} imagens baixadas",
    }))
    out_q.put(("DONE", None))


