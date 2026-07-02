import queue
import threading


def run_extrator(url: str, output_path: str, out_q: queue.Queue, stop_ev: threading.Event):
    def log(msg, tag="normal"):
        out_q.put((tag, msg + "\n"))

    try:
        import requests
        from bs4 import BeautifulSoup
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        log(f"  [ERRO] Dependencia nao instalada: {e}", "error")
        out_q.put(("DONE", None))
        return

    if not url.strip():
        log("  [ERRO] Nenhuma URL informada.", "error")
        out_q.put(("DONE", None))
        return

    if not url.startswith("http"):
        url = "https://" + url

    log(f"  Acessando: {url}", "info")
    out_q.put(("PROGRESS", 0.1))

    try:
        r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
    except Exception as e:
        log(f"  [ERRO] Nao foi possivel acessar a URL: {e}", "error")
        out_q.put(("DONE", None))
        return

    if stop_ev.is_set():
        out_q.put(("DONE", None))
        return

    soup = BeautifulSoup(r.text, "html.parser")
    tabelas = soup.find_all("table")
    log(f"  {len(tabelas)} tabela(s) encontrada(s) na pagina.", "info")

    if not tabelas:
        log("  Nenhuma tabela HTML encontrada nessa pagina.", "warning")
        out_q.put(("HISTORY", {"modulo": "extrator", "status": "sem_tabelas",
                               "detalhes": f"0 tabelas em {url}"}))
        out_q.put(("DONE", None))
        return

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for idx, tabela in enumerate(tabelas, 1):
            if stop_ev.is_set():
                break
            out_q.put(("PROGRESS", idx / len(tabelas)))
            try:
                df = pd.read_html(str(tabela))[0]
                sheet_name = f"Tabela_{idx}"
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                log(f"  Tabela {idx}: {df.shape[0]} linhas x {df.shape[1]} colunas  -> {sheet_name}", "success")
            except Exception as e:
                log(f"  Tabela {idx}: erro ao processar ({e})", "warning")

    # Formata cabecalhos
    try:
        wb = load_workbook(output_path)
        fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                vals = [str(c.value) for c in col if c.value is not None]
                if vals:
                    ws.column_dimensions[col[0].column_letter].width = min(max(len(v) for v in vals) + 2, 50)
        wb.save(output_path)
    except Exception:
        pass

    log("")
    log("  " + "=" * 56, "dim")
    log("  EXTRACAO FINALIZADA!", "title")
    log(f"  {len(tabelas)} tabela(s) exportada(s) para:", "success")
    log(f"  {output_path}", "info")
    log("  " + "=" * 56, "dim")

    out_q.put(("HISTORY", {
        "modulo": "extrator",
        "status": "sucesso",
        "detalhes": f"{len(tabelas)} tabelas extraidas de {url}",
    }))
    out_q.put(("DONE", None))
