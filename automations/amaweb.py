import os
import json
import time
import math
import queue
import threading
from datetime import datetime


TIMEOUT = 180
_HIST_FILE = "historico_amaweb.json"


def _load_hist_ama() -> dict:
    try:
        with open(_HIST_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_hist_ama(data: dict) -> None:
    with open(_HIST_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def run_amaweb(urls_path: str, resultado_path: str, out_q: queue.Queue,
               stop_ev: threading.Event, threshold: float = 0.0):
    def log(msg, tag="normal"):
        out_q.put((tag, msg + "\n"))

    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        log(f"  [ERRO] Dependencia nao instalada: {e}", "error")
        log("  Execute: pip install selenium openpyxl", "warning")
        out_q.put(("DONE", None))
        return

    if not os.path.exists(urls_path):
        log(f"  [ERRO] Arquivo nao encontrado: {urls_path}", "error")
        out_q.put(("DONE", None))
        return

    with open(urls_path, "r", encoding="utf-8") as f:
        urls = [l.strip() for l in f if l.strip()]

    total = len(urls)
    log(f"  Iniciando avaliacao de acessibilidade em {total} site(s)...", "info")
    log(f"  Timeout por site: {TIMEOUT}s", "info")
    log(f"  Resultado: {resultado_path}", "info")
    if threshold > 0:
        log(f"  Alerta de nota baixa: sites abaixo de {threshold}", "warning")
    log("")

    hist_ama = _load_hist_ama()
    agora_run = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Cabecalho com data da execucao
    ws.append(["URL", "Nota", "Execucao"])
    hdr_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")

    try:
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        log(f"  [ERRO] ChromeDriver nao iniciado: {e}", "error")
        log("  Verifique se o Google Chrome esta instalado.", "warning")
        out_q.put(("DONE", None))
        return

    inicio_total = time.time()
    notas_validas: list[float] = []
    alertas: list[str] = []
    pares_resultado: list[tuple[str, str]] = []  # (site, nota) para download TXT

    try:
        for i, site in enumerate(urls, 1):
            if stop_ev.is_set():
                log("\n  [!] Execucao interrompida.", "warning")
                break

            url = (
                f"https://amaweb.unifesp.br/avaliador/results/{site}"
                if not site.startswith("http") else site
            )
            log(f"  [{i:03d}/{total}]  Avaliando -> {site}", "info")

            try:
                driver.get(url)
                nota_el = WebDriverWait(driver, TIMEOUT).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".score-circle-value"))
                )
                nota = nota_el.text.strip().replace("Nota:", "").strip() or "Erro"
            except Exception:
                nota = "Timeout / Erro"

            # Converte nota para float para comparacao com threshold
            nota_f = None
            try:
                nota_f = float(nota.replace(",", "."))
                notas_validas.append(nota_f)
            except ValueError:
                pass

            # Salva no historico por site
            if site not in hist_ama:
                hist_ama[site] = []
            hist_ama[site].append({"data": agora_run, "nota": nota})
            hist_ama[site] = hist_ama[site][-50:]  # mantém últimas 50

            ws.append([site, nota, agora_run])
            pares_resultado.append((site, nota))

            # Destaca linha se abaixo do threshold
            row_idx = ws.max_row
            if nota_f is not None and threshold > 0 and nota_f < threshold:
                alerta_fill = PatternFill(start_color="F85149", end_color="F85149", fill_type="solid")
                for cell in ws[row_idx]:
                    cell.fill = alerta_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                alertas.append(f"{site} ({nota})")

            wb.save(resultado_path)

            decorrido = time.time() - inicio_total
            media     = decorrido / i
            restante  = (total - i) * media

            is_erro = nota in ("Erro", "Timeout / Erro")
            is_baixo = nota_f is not None and threshold > 0 and nota_f < threshold
            if is_baixo:
                tag = "error"
            elif is_erro:
                tag = "warning"
            else:
                tag = "success"

            log(
                f"           Nota: {nota:<10}  media: {media:.1f}s/site"
                f"  restante: ~{restante / 60:.1f}min",
                tag,
            )

            pct = i / total
            barra = "X" * int(30 * pct) + "." * (30 - int(30 * pct))
            out_q.put(("progress", f"  [{barra}] {math.floor(pct * 100)}%\n"))
            out_q.put(("PROGRESS", pct))

    finally:
        driver.quit()
        wb.save(resultado_path)
        _save_hist_ama(hist_ama)

    # Ajusta largura das colunas
    try:
        from openpyxl import load_workbook as _lw
        _wb2 = _lw(resultado_path)
        _ws2 = _wb2.active
        for col in _ws2.columns:
            vals = [str(c.value or "") for c in col]
            _ws2.column_dimensions[col[0].column_letter].width = min(max(len(v) for v in vals) + 2, 60)
        _wb2.save(resultado_path)
    except Exception:
        pass

    media_geral = sum(notas_validas) / len(notas_validas) if notas_validas else 0.0

    log("")
    log("  " + "=" * 56, "dim")
    log("  AVALIACAO FINALIZADA!", "title")
    log(f"  Sites avaliados : {total}", "info")
    log(f"  Media geral     : {media_geral:.2f}", "success" if media_geral >= (threshold or 5) else "warning")
    if alertas:
        log(f"  Abaixo de {threshold}: {len(alertas)} site(s)", "error")
        for a in alertas:
            log(f"    -> {a}", "error")
    log(f"  Planilha salva  : {resultado_path}", "success")
    log("  " + "=" * 56, "dim")

    out_q.put(("PROGRESS", 1.0))
    out_q.put(("RESULT_AMA", pares_resultado))

    out_q.put(("HISTORY", {
        "modulo": "amaweb",
        "status": "sucesso",
        "detalhes": (
            f"{total} sites | media {media_geral:.2f}"
            + (f" | {len(alertas)} abaixo de {threshold}" if alertas else "")
        ),
    }))
    out_q.put(("DONE", None))
