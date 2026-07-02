import os
import queue
import threading
from datetime import datetime


def run_relatorio(output_path: str, out_q: queue.Queue, stop_ev: threading.Event):
    def log(msg, tag="normal"):
        out_q.put((tag, msg + "\n"))

    try:
        import pandas as pd
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        log(f"  [ERRO] Dependencia nao instalada: {e}", "error")
        out_q.put(("DONE", None))
        return

    log("  Gerando relatorio consolidado...", "info")
    out_q.put(("PROGRESS", 0.05))

    wb = Workbook()
    sheets_criadas = []

    # ── Helpers de estilo ──────────────────────────────────────────────
    HDR_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    ALT_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(bold=True, size=14, color="1F3864")
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _estilizar_sheet(ws):
        for cell in ws[1]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
        for col in ws.columns:
            vals = [str(c.value or "") for c in col]
            width = min(max((len(v) for v in vals), default=10) + 2, 60)
            ws.column_dimensions[col[0].column_letter].width = width
        ws.row_dimensions[1].height = 22

    # ── Sheet 1: Historico de execucoes ────────────────────────────────
    out_q.put(("PROGRESS", 0.2))
    try:
        import json
        hist_path = os.path.join(os.path.dirname(__file__), "..", "historico.json")
        with open(hist_path, "r", encoding="utf-8") as f:
            historico = json.load(f)

        ws_hist = wb.active
        ws_hist.title = "Historico"
        ws_hist.append(["Modulo", "Data", "Status", "Detalhes"])
        for entry in historico:
            ws_hist.append([
                entry.get("modulo", ""),
                entry.get("data", ""),
                entry.get("status", ""),
                entry.get("detalhes", ""),
            ])
        _estilizar_sheet(ws_hist)
        sheets_criadas.append("Historico")
        log(f"  Historico: {len(historico)} registros", "success")
    except Exception as e:
        log(f"  Historico: nao encontrado ({e})", "warning")

    if stop_ev.is_set():
        out_q.put(("DONE", None))
        return

    # ── Sheet 2: Dados do Scraper ──────────────────────────────────────
    out_q.put(("PROGRESS", 0.45))
    scraper_xlsx = os.path.join("galeria_noticias", "registro_noticias.xlsx")
    if os.path.exists(scraper_xlsx):
        try:
            df_sc = pd.read_excel(scraper_xlsx)
            ws_sc = wb.create_sheet("Scraper")
            ws_sc.append(list(df_sc.columns))
            for _, row in df_sc.iterrows():
                ws_sc.append(list(row))
            _estilizar_sheet(ws_sc)
            sheets_criadas.append("Scraper")
            log(f"  Scraper: {len(df_sc)} registros de imagens", "success")
        except Exception as e:
            log(f"  Scraper: erro ao ler planilha ({e})", "warning")
    else:
        log("  Scraper: registro_noticias.xlsx nao encontrado", "dim")

    if stop_ev.is_set():
        out_q.put(("DONE", None))
        return

    # ── Sheet 3: Dados do AMAWeb ───────────────────────────────────────
    out_q.put(("PROGRESS", 0.65))
    ama_xlsx = "resultado.xlsx"
    if os.path.exists(ama_xlsx):
        try:
            df_ama = pd.read_excel(ama_xlsx)
            ws_ama = wb.create_sheet("AMAWeb")
            ws_ama.append(list(df_ama.columns))
            for _, row in df_ama.iterrows():
                ws_ama.append(list(row))
            _estilizar_sheet(ws_ama)
            sheets_criadas.append("AMAWeb")
            log(f"  AMAWeb: {len(df_ama)} sites avaliados", "success")
        except Exception as e:
            log(f"  AMAWeb: erro ao ler planilha ({e})", "warning")
    else:
        log("  AMAWeb: resultado.xlsx nao encontrado", "dim")

    if stop_ev.is_set():
        out_q.put(("DONE", None))
        return

    # ── Sheet 4: Sumario ───────────────────────────────────────────────
    out_q.put(("PROGRESS", 0.85))
    ws_sum = wb.create_sheet("Sumario", 0)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 40

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws_sum["A1"] = "AutoHub Pro — Relatorio Consolidado"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum.merge_cells("A1:B1")
    ws_sum.row_dimensions[1].height = 30

    dados = [
        ("Gerado em", now_str),
        ("Sheets incluidas", ", ".join(sheets_criadas)),
        ("", ""),
        ("Arquivo", os.path.basename(output_path)),
    ]
    for i, (k, v) in enumerate(dados, 3):
        ws_sum[f"A{i}"] = k
        ws_sum[f"B{i}"] = v
        ws_sum[f"A{i}"].font = Font(bold=True)

    sheets_criadas.insert(0, "Sumario")

    wb.save(output_path)
    out_q.put(("PROGRESS", 1.0))

    log("")
    log("  " + "=" * 56, "dim")
    log("  RELATORIO GERADO COM SUCESSO!", "title")
    log(f"  Sheets: {', '.join(sheets_criadas)}", "success")
    log(f"  Arquivo: {output_path}", "info")
    log("  " + "=" * 56, "dim")

    out_q.put(("HISTORY", {
        "modulo": "relatorio",
        "status": "sucesso",
        "detalhes": f"{len(sheets_criadas)} sheets | {output_path}",
    }))
    out_q.put(("DONE", None))
